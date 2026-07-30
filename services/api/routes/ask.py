"""Ask surface — AI assistant grounded in connected data, with memory and file processing."""

from __future__ import annotations

import base64
import time
import uuid
from collections import Counter
from datetime import datetime, timezone
from typing import Any

import httpx
from fastapi import APIRouter, File, Form, UploadFile
from pydantic import BaseModel, Field

from libs.config.settings import get_settings
from libs.observability.logging import get_logger
from services.api.routes.ingest import EntityRecord, get_entity_store
from services.api.routes.memory import (
    add_memory,
    extract_memories_from_conversation,
    get_memories,
)

logger = get_logger("ask")

router = APIRouter()


class ChatMessage(BaseModel):
    role: str = "user"
    content: str


class AskRequest(BaseModel):
    question: str = Field(..., min_length=1, max_length=4000)
    viewer_id: uuid.UUID = uuid.UUID("00000000-0000-0000-0000-000000000001")
    history: list[ChatMessage] = []
    file_contents: list[dict[str, str]] = []


class Citation(BaseModel):
    source: str
    entity_name: str
    entity_type: str
    snippet: str


class AnswerEnvelope(BaseModel):
    answer: str
    citations: list[Citation]
    freshness: str
    layer: str
    conflicts: list[str]
    latency_ms: int


def _get_llm_config(settings) -> tuple[str, dict[str, str]]:
    """Get LLM endpoint and headers, routing through Portkey if available."""
    openai_key = settings.openai_api_key.get_secret_value()
    portkey_key = settings.portkey_api_key.get_secret_value()

    headers: dict[str, str] = {"Content-Type": "application/json"}

    # Portkey requires a saved virtual-key integration (not inline provider).
    # Use direct OpenAI until a Portkey virtual key is configured.
    # To enable: create a virtual key in Portkey dashboard, then set
    # x-portkey-virtual-key header instead of Authorization.
    headers["Authorization"] = f"Bearer {openai_key}"
    base_url = "https://api.openai.com/v1"

    if portkey_key:
        logger.debug("llm_routing_note", msg="Portkey key present but using direct OpenAI (needs virtual-key config)")

    return base_url, headers


async def _get_connected_sources(settings) -> list[dict]:
    """Check Nango for real connected sources."""
    secret = settings.nango_secret_key.get_secret_value()
    if not secret:
        return []
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                f"{settings.nango_base_url}/connections",
                headers={"Authorization": f"Bearer {secret}"},
                timeout=8.0,
            )
            if resp.status_code == 200:
                data = resp.json()
                connections = (
                    data.get("connections", [])
                    if isinstance(data, dict)
                    else data
                )
                seen = {}
                for c in connections:
                    p = c.get(
                        "provider_config_key", c.get("provider", "unknown")
                    )
                    if p not in seen:
                        seen[p] = p.replace("-", " ").title()
                return [{"provider": k, "name": v} for k, v in seen.items()]
    except Exception:
        pass
    return []


def _get_canon_context() -> str:
    """Get relevant canon assertions for the system prompt."""
    try:
        from services.api.routes.canon import _assertions, AssertionStatus

        active = [a for a in _assertions if a.status == AssertionStatus.ACTIVE]
        if not active:
            return ""

        lines = ["Company knowledge (Canon):"]
        for a in active[:20]:
            lines.append(
                f"  - {a.entity_name} ({a.entity_type}): {a.field} = {a.value} "
                f"[source: {a.source}, by: {a.author}]"
            )
        return "\n".join(lines)
    except Exception:
        return ""


def _parse_email_date(date_str: str) -> datetime:
    """Parse RFC 2822 email dates into datetime objects for proper sorting."""
    from email.utils import parsedate_to_datetime

    if not date_str:
        return datetime(2000, 1, 1, tzinfo=timezone.utc)
    try:
        return parsedate_to_datetime(date_str)
    except Exception:
        pass
    # Fallback: try common formats
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%a, %d %b %Y %H:%M:%S %z"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except Exception:
            continue
    return datetime(2000, 1, 1, tzinfo=timezone.utc)


def _build_data_analytics(store: list[EntityRecord]) -> str:
    """Pre-compute analytics from the entity store so the LLM can answer
    data-driven questions (top contacts, entity counts, timelines, etc.)."""
    if not store:
        return ""

    # Group by type
    by_type: dict[str, list[EntityRecord]] = {}
    for e in store:
        by_type.setdefault(e.type, []).append(e)

    lines: list[str] = []
    lines.append(f"DATA SUMMARY ({len(store)} total entities):")
    for t, entities in sorted(by_type.items(), key=lambda x: -len(x[1])):
        lines.append(f"  {t}: {len(entities)}")

    # Email analytics
    emails = by_type.get("email", [])
    if emails:
        lines.append(f"\nEMAIL ANALYTICS ({len(emails)} emails):")

        # Sort emails by date (newest first) using proper date parsing
        sorted_emails = sorted(
            emails,
            key=lambda e: _parse_email_date(e.properties.get("date", "")),
            reverse=True,
        )

        # Recent email timeline (so LLM can answer "what's my recent email")
        lines.append("  Recent emails (newest first):")
        for e in sorted_emails[:15]:
            from_addr = e.properties.get("from", "")
            date_str = e.properties.get("date", "")
            snippet = e.properties.get("snippet", "")[:100]
            labels = e.properties.get("labels", "")
            lines.append(
                f"    [{date_str}] \"{e.name}\" from {from_addr}"
                + (f" — {snippet}" if snippet else "")
                + (f" (labels: {labels})" if labels else "")
            )

        # Top senders — filter out system/notification senders
        JUNK_SENDERS = {
            "noreply", "no-reply", "notification", "mailer-daemon",
            "calendar-notification", "drive-shares-dm-noreply",
        }

        def _is_junk_sender(addr: str) -> bool:
            addr_lower = addr.lower()
            return any(j in addr_lower for j in JUNK_SENDERS) or not addr.strip()

        sender_counts: Counter[str] = Counter()
        recipient_counts: Counter[str] = Counter()
        for e in emails:
            from_addr = e.properties.get("from", "").strip()
            to_addr = e.properties.get("to", "")
            if from_addr and not _is_junk_sender(from_addr):
                sender_counts[from_addr] += 1
            if to_addr:
                for addr in to_addr.split(","):
                    addr = addr.strip()
                    if addr and not _is_junk_sender(addr):
                        recipient_counts[addr] += 1

        lines.append("  Top senders (by email count):")
        for sender, count in sender_counts.most_common(15):
            lines.append(f"    {sender}: {count} emails")

        lines.append("  Top recipients:")
        for recip, count in recipient_counts.most_common(15):
            lines.append(f"    {recip}: {count} emails")

        # Contact frequency (combined send + receive)
        all_contacts: Counter[str] = Counter()
        all_contacts.update(sender_counts)
        all_contacts.update(recipient_counts)
        lines.append("  Top contacts (by total email volume):")
        for contact, count in all_contacts.most_common(15):
            lines.append(f"    {contact}: {count} emails")

        # Labels distribution
        label_counts: Counter[str] = Counter()
        for e in emails:
            labels = e.properties.get("labels", "")
            for lbl in labels.split(", "):
                if lbl.strip():
                    label_counts[lbl.strip()] += 1
        if label_counts:
            lines.append("  Email labels:")
            for lbl, cnt in label_counts.most_common(10):
                lines.append(f"    {lbl}: {cnt}")

    # People analytics
    people = by_type.get("person", [])
    if people:
        lines.append(f"\nPEOPLE ({len(people)} contacts):")
        # Group by source
        by_source: Counter[str] = Counter()
        for p in people:
            by_source[p.source] += 1
        for src, cnt in by_source.most_common():
            lines.append(f"  From {src}: {cnt}")

        # Show all people with their details
        lines.append("  Contacts list:")
        for p in people[:30]:
            email = p.properties.get("email", "")
            company = p.properties.get("company", "")
            extras = []
            if email:
                extras.append(email)
            if company:
                extras.append(f"at {company}")
            detail = f" ({', '.join(extras)})" if extras else ""
            lines.append(f"    {p.name}{detail} [source: {p.source}]")

    # Company analytics
    companies = by_type.get("company", [])
    if companies:
        lines.append(f"\nCOMPANIES ({len(companies)}):")
        for c in companies[:20]:
            domain = c.properties.get("domain", "")
            contacts = c.properties.get("contact_count", "")
            extras = []
            if domain:
                extras.append(domain)
            if contacts:
                extras.append(f"{contacts} contacts")
            detail = f" ({', '.join(str(x) for x in extras)})" if extras else ""
            lines.append(f"  {c.name}{detail} [source: {c.source}]")

    # Deal analytics
    deals = by_type.get("deal", [])
    if deals:
        lines.append(f"\nDEALS ({len(deals)}):")
        for d in deals[:20]:
            amount = d.properties.get("amount", "")
            stage = d.properties.get("stage", "")
            extras = []
            if amount:
                extras.append(f"${amount}")
            if stage:
                extras.append(stage)
            detail = f" ({', '.join(extras)})" if extras else ""
            lines.append(f"  {d.name}{detail} [source: {d.source}]")

    # Document analytics
    docs = by_type.get("document", [])
    if docs:
        lines.append(f"\nDOCUMENTS ({len(docs)}):")
        for d in docs[:15]:
            mime = d.properties.get("mime_type", "")
            owner = d.properties.get("owner", "")
            modified = d.properties.get("modified", "")[:10]
            extras = []
            if mime:
                extras.append(mime.split("/")[-1])
            if owner:
                extras.append(f"by {owner}")
            if modified:
                extras.append(modified)
            detail = f" ({', '.join(extras)})" if extras else ""
            lines.append(f"  {d.name}{detail}")

    # Spreadsheet analytics
    sheets = by_type.get("spreadsheet", [])
    if sheets:
        lines.append(f"\nSPREADSHEETS ({len(sheets)}):")
        for s in sheets[:10]:
            owner = s.properties.get("owner", "")
            modified = s.properties.get("modified", "")[:10]
            extras = []
            if owner:
                extras.append(f"by {owner}")
            if modified:
                extras.append(modified)
            detail = f" ({', '.join(extras)})" if extras else ""
            lines.append(f"  {s.name}{detail}")

    return "\n".join(lines)


def _build_relevant_context(
    store: list[EntityRecord], question: str
) -> tuple[list[str], list[dict]]:
    """Build relevant entity context lines for the question."""
    context_lines: list[str] = []
    relevant_entities: list[dict] = []

    if not store:
        return context_lines, relevant_entities

    q = question.lower()

    # Detect if user is asking about recency / timeline
    recency_keywords = {"recent", "latest", "last", "newest", "new", "today", "yesterday"}
    is_recency_query = any(kw in q for kw in recency_keywords)

    matched: list[EntityRecord] = []
    for entity in store:
        name_match = any(
            word in entity.name.lower()
            for word in q.split()
            if len(word) > 2
        )
        type_match = entity.type in q
        if name_match or type_match:
            matched.append(entity)

    # Sort by date if recency query and entities have dates
    if is_recency_query and matched:
        matched.sort(
            key=lambda e: _parse_email_date(
                e.properties.get("date", e.fetched_at or "")
            ),
            reverse=True,
        )

    for entity in matched[:30]:
        props_str = ""
        if entity.properties:
            top_props = {
                k: str(v)[:150]
                for k, v in list(entity.properties.items())[:8]
                if not k.startswith("_")
            }
            props_str = f" | {top_props}" if top_props else ""

        context_lines.append(
            f"[{entity.source}] {entity.type.title()}: {entity.name}{props_str}"
        )
        relevant_entities.append(
            {
                "source": entity.source,
                "name": entity.name,
                "type": entity.type,
            }
        )

    return context_lines, relevant_entities


@router.post("/ask/upload")
async def ask_with_files(
    question: str = Form(...),
    viewer_id: str = Form("00000000-0000-0000-0000-000000000001"),
    history: str = Form("[]"),
    files: list[UploadFile] = File(default=[]),
) -> AnswerEnvelope:
    """Handle questions with file attachments (images, PDFs, docs)."""
    import json

    parsed_history = []
    try:
        parsed_history = [ChatMessage(**m) for m in json.loads(history)]
    except Exception:
        pass

    file_contents: list[dict[str, str]] = []
    for f in files:
        content = await f.read()
        mime = f.content_type or ""
        fname = f.filename or "file"

        if mime.startswith("image/"):
            b64 = base64.b64encode(content).decode("utf-8")
            file_contents.append({
                "name": fname,
                "type": "image",
                "content": f"data:{mime};base64,{b64}",
                "mime": mime,
            })
        else:
            text = ""

            # PDF extraction
            if mime == "application/pdf" or fname.lower().endswith(".pdf"):
                try:
                    import io
                    from PyPDF2 import PdfReader
                    reader = PdfReader(io.BytesIO(content))
                    pages = []
                    for i, page in enumerate(reader.pages):
                        page_text = page.extract_text() or ""
                        if page_text.strip():
                            pages.append(f"[Page {i+1}]\n{page_text}")
                    text = "\n\n".join(pages)
                    logger.info("pdf_extracted", filename=fname, pages=len(reader.pages), chars=len(text))
                except Exception as e:
                    logger.warning("pdf_extraction_failed", filename=fname, error=str(e))
                    text = f"[Could not extract text from PDF: {fname}]"

            # DOCX extraction
            elif fname.lower().endswith(".docx") or mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                try:
                    import io
                    from docx import Document as DocxDocument
                    doc = DocxDocument(io.BytesIO(content))
                    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
                    logger.info("docx_extracted", filename=fname, chars=len(text))
                except Exception as e:
                    logger.warning("docx_extraction_failed", filename=fname, error=str(e))
                    text = f"[Could not extract text from DOCX: {fname}]"

            # Excel (.xlsx) extraction
            elif fname.lower().endswith(".xlsx") or mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                try:
                    import io
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                    sheets_text = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        rows = []
                        for row in ws.iter_rows(values_only=True):
                            cells = [str(c) if c is not None else "" for c in row]
                            if any(cells):
                                rows.append(" | ".join(cells))
                        if rows:
                            sheets_text.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
                    text = "\n\n".join(sheets_text)
                    wb.close()
                    logger.info("xlsx_extracted", filename=fname, sheets=len(wb.sheetnames), chars=len(text))
                except Exception as e:
                    logger.warning("xlsx_extraction_failed", filename=fname, error=str(e))
                    text = f"[Could not extract data from Excel file: {fname}]"

            # Plain text / CSV / other text formats
            else:
                try:
                    text = content.decode("utf-8", errors="replace")
                except Exception:
                    text = content.decode("latin-1", errors="replace")

            if not text.strip():
                text = f"[No readable text extracted from {fname}]"

            if len(text) > 15000:
                text = text[:15000] + "\n... [truncated]"

            file_contents.append({
                "name": fname,
                "type": "document",
                "content": text,
                "mime": mime,
            })

    request = AskRequest(
        question=question,
        viewer_id=uuid.UUID(viewer_id),
        history=parsed_history,
        file_contents=file_contents,
    )
    return await ask_question(request)


@router.post("/ask", response_model=AnswerEnvelope)
async def ask_question(request: AskRequest) -> AnswerEnvelope:
    settings = get_settings()
    start = time.time()

    openai_key = settings.openai_api_key.get_secret_value()
    store = get_entity_store()
    connected_sources = await _get_connected_sources(settings)

    if not openai_key:
        return AnswerEnvelope(
            answer="OpenAI API key is not configured. Set OPENAI_API_KEY in your .env file.",
            citations=[],
            freshness="N/A",
            layer="system",
            conflicts=[],
            latency_ms=int((time.time() - start) * 1000),
        )

    base_url, headers = _get_llm_config(settings)

    # Build pre-computed analytics from ALL entity data
    analytics_context = _build_data_analytics(store)

    # Build keyword-relevant entity context
    context_lines, relevant_entities = _build_relevant_context(
        store, request.question
    )

    # Get memory context
    user_id = str(request.viewer_id)
    memories = get_memories(user_id, limit=15)
    memory_context = ""
    if memories:
        memory_lines = [
            f"  - [{m.category}] {m.content}" for m in memories[:15]
        ]
        memory_context = (
            "\n\nThings I remember about this user:\n"
            + "\n".join(memory_lines)
        )

    # Get canon context
    canon_context = _get_canon_context()

    # Build system prompt
    source_names = [s["name"] for s in connected_sources]
    sources_str = ", ".join(source_names) if source_names else "none"

    base_instructions = (
        "You are the Optimus TrustLayer Knowledge Assistant — an enterprise-grade AI "
        "that helps RevOps and CS teams understand their data across connected tools.\n\n"
        "RULES:\n"
        "- Answer using the connected data analytics and entity details below\n"
        "- When asked about top contacts, email volume, etc., use the pre-computed analytics\n"
        "- Always give specific names, numbers, and rankings from the data\n"
        "- Be concise, professional, and cite sources when referencing data\n"
        "- Format with markdown: use **bold** for names, numbered lists for rankings\n"
        "- When analyzing uploaded files, describe what you see and relate to the data\n"
        "- Reference company canon assertions when relevant\n"
        "- Never say 'the data doesn't provide' if the analytics section has the answer\n"
    )

    if store:
        system_prompt = (
            base_instructions
            + f"\nConnected sources: {sources_str}\n\n"
            + analytics_context
        )
        if context_lines:
            system_prompt += (
                f"\n\nAdditional matching entities:\n"
                + "\n".join(context_lines[:20])
            )
    elif connected_sources:
        system_prompt = (
            base_instructions
            + f"\nThe user has connected: {sources_str}. "
            "Data ingestion may not have run yet.\n"
            "Guide them to run ingestion from Sources page or try again shortly.\n"
        )
    else:
        system_prompt = (
            base_instructions
            + "\nNo data sources connected yet. Guide the user to connect tools via Sources.\n"
        )

    if canon_context:
        system_prompt += f"\n\n{canon_context}"

    if memory_context:
        system_prompt += memory_context

    # Build messages
    messages: list[dict[str, Any]] = [
        {"role": "system", "content": system_prompt}
    ]

    for msg in request.history[-10:]:
        messages.append({"role": msg.role, "content": msg.content})

    # Build user message (potentially multimodal)
    has_images = any(
        fc.get("type") == "image" for fc in request.file_contents
    )

    if has_images or request.file_contents:
        content_parts: list[dict[str, Any]] = []

        file_text_parts = []
        for fc in request.file_contents:
            if fc["type"] == "document":
                file_text_parts.append(
                    f"\n--- File: {fc['name']} ---\n{fc['content']}\n---"
                )

        if file_text_parts:
            content_parts.append({
                "type": "text",
                "text": request.question + "\n\nAttached files:" + "".join(file_text_parts),
            })
        else:
            content_parts.append({
                "type": "text",
                "text": request.question,
            })

        for fc in request.file_contents:
            if fc["type"] == "image":
                content_parts.append({
                    "type": "image_url",
                    "image_url": {"url": fc["content"], "detail": "auto"},
                })

        messages.append({"role": "user", "content": content_parts})
        model = "gpt-4o"
    else:
        messages.append({"role": "user", "content": request.question})
        model = "gpt-4o-mini"

    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                f"{base_url}/chat/completions",
                headers=headers,
                json={
                    "model": model,
                    "messages": messages,
                    "temperature": 0.3,
                    "max_tokens": 2000,
                },
                timeout=60.0,
            )

            if resp.status_code != 200:
                error_detail = (
                    resp.json().get("error", {}).get("message", resp.text[:200])
                )
                return AnswerEnvelope(
                    answer=f"AI service error: {error_detail}",
                    citations=[],
                    freshness=datetime.now(timezone.utc).isoformat(),
                    layer="system",
                    conflicts=[],
                    latency_ms=int((time.time() - start) * 1000),
                )

            completion = resp.json()
            answer_text = completion["choices"][0]["message"]["content"]

    except Exception as e:
        logger.error("ask_llm_error", error=str(e))
        answer_text = f"Error generating answer: {str(e)}"

    # Build citations
    cited: list[Citation] = []

    # If files were uploaded, cite the uploaded files first
    if request.file_contents:
        for fc in request.file_contents:
            cited.append(
                Citation(
                    source="uploaded-file",
                    entity_name=fc.get("name", "Uploaded file"),
                    entity_type="document" if fc.get("type") == "document" else "image",
                    snippet=f"Uploaded by user",
                )
            )

    # Then cite entities actually mentioned in the answer
    answer_lower = answer_text.lower()
    for ent in relevant_entities:
        if ent["name"].lower() in answer_lower:
            cited.append(
                Citation(
                    source=ent["source"],
                    entity_name=ent["name"],
                    entity_type=ent["type"],
                    snippet=f"From {ent['source']}",
                )
            )
            if len(cited) >= 8:
                break

    # If no file citations and no entity matches, cite data sources
    if not cited and store:
        source_set = {e.source for e in store}
        cited = [
            Citation(
                source=src,
                entity_name=f"{sum(1 for e in store if e.source == src)} entities",
                entity_type="summary",
                snippet=f"Aggregated from {src}",
            )
            for src in list(source_set)[:5]
        ]

    citations = cited

    layer = "personal" if store else "system"
    if canon_context:
        layer = "company"

    latency = int((time.time() - start) * 1000)

    # Background: extract memories from this conversation
    try:
        import asyncio

        asyncio.create_task(
            extract_memories_from_conversation(
                user_id, request.question, answer_text
            )
        )
    except Exception:
        pass

    return AnswerEnvelope(
        answer=answer_text,
        citations=citations,
        freshness=datetime.now(timezone.utc).isoformat(),
        layer=layer,
        conflicts=[],
        latency_ms=latency,
    )
